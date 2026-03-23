import logging
import os
import tempfile
import pandas as pd
from aiogram import Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.types import FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton
from sqlalchemy import func
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from database import get_db_session
from models import AdminAccount
from .states import AdminStates
from .utils import check_admin
from .keyboards import get_admin_keyboard

logger = logging.getLogger(__name__)
router = Router()

back_keyboard = InlineKeyboardMarkup(
    inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ Back", callback_data="admin_back")]
    ]
)

@router.message(F.text == "📤 Export Data")
async def handle_export_data(message: types.Message, state: FSMContext):
    """Handle export data button"""
    if not await check_admin(message):
        return

    await message.answer(
        "📤 **Export Admin Data**\n\n"
        "Please send the date for export.\n"
        "Format: `YYYY-MM-DD`\n\n"
        "Example: `2026-02-07`",
        parse_mode='Markdown',
        reply_markup=back_keyboard
    )
    await state.set_state(AdminStates.waiting_for_export_date)


@router.message(AdminStates.waiting_for_export_date)
async def process_export_date(message: types.Message, state: FSMContext):
    """Process export date and create Excel with custom filename"""
    session = get_db_session()
    try:
        date_str = message.text.strip()
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            await message.answer(
                "❌ Invalid date format.\n"
                "Please use: `YYYY-MM-DD`",
                parse_mode='Markdown',
                reply_markup=back_keyboard
            )
            return

        accounts = session.query(AdminAccount).filter(
            func.date(AdminAccount.entry_date) == target_date
        ).all()

        if not accounts:
            await message.answer(
                f"❌ No admin data found for `{date_str}`",
                parse_mode='Markdown',
                reply_markup=back_keyboard
            )
            return

        data = [[acc.username, acc.password] for acc in accounts]
        df = pd.DataFrame(data, columns=['username', 'password'])
        count = len(df)

        download_name = f"{count}_accounts_{date_str}.xlsx"

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_filepath = tmp.name

        with pd.ExcelWriter(temp_filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Accounts')

            workbook = writer.book
            worksheet = writer.sheets['Accounts']

            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            data_font = Font(name='Segoe UI', size=10)
            data_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for i, column in enumerate(worksheet.columns):
                max_length = 0
                column_letter = get_column_letter(i + 1)
                for cell in column:
                    if cell.row > 1:
                        cell.font = data_font
                        cell.alignment = data_alignment
                        cell.border = thin_border
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                adjusted_width = min((max_length + 4) * 1.1, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            worksheet.freeze_panes = 'A2'

        await message.answer_document(
            FSInputFile(path=temp_filepath, filename=download_name),
            caption=f"📤 Exported `{count}` accounts for `{date_str}`",
            parse_mode='Markdown'
        )

        await state.clear()

        if os.path.exists(temp_filepath):
            os.unlink(temp_filepath)

        logger.info(f"Exported {count} accounts for {date_str} as {download_name}")

    except Exception as e:
        logger.error(f"Error exporting data: {e}", exc_info=True)
        await message.answer(f"❌ Error exporting data: {e}")
        await state.clear()
    finally:
        session.close()


@router.callback_query(F.data == "admin_back")
async def process_back_button(callback: types.CallbackQuery, state: FSMContext):
    """Handle back button press"""
    await state.clear()
    try:
        await callback.message.delete()
    except Exception:
        pass

    await callback.message.answer(
        "🛍 **Admin Panel**",
        parse_mode='Markdown',
        reply_markup=get_admin_keyboard()
    )
    await callback.answer()